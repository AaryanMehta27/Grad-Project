"""
06_analyze_validation.py
========================
Tier-2 human-validity analysis: compares the human ratings (blind workbooks in
data/processed/validation/) against the LLM scores (validation_sample_100_KEY.csv).

Computes, per Section 3.4 / Decision 5.27 of the methodology:
  - Cohen's kappa (unweighted + linear-weighted) LLM vs each rater, per criterion
  - Raw percent agreement and confusion matrices (prevalence/kappa guard)
  - Pooled kappa across all four criteria (overall headline)
  - Consensus-subset kappa (tasks where the human raters agree)
  - Inter-rater: Cohen's kappa + Krippendorff's alpha (ordinal) between raters
  - Composite comparison: Design-Y composite from each rater's scores vs the
    LLM composite — gate (0 vs >0) agreement/kappa and Spearman rho
  - Flagged tasks excluded from kappa per protocol (none were flagged)
  - Disagreement listing for qualitative review

Run after the rating workbooks are filled. Read-only; prints a report.
Author: Aaryan Mehta. Created 06 June 2026.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import Counter, defaultdict
from pathlib import Path
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
VAL = ROOT / "data" / "processed" / "validation"
CRIT = ["IS", "OV", "CtxI", "CM"]
CRIT_FULL = {"IS": "Information Sufficiency", "OV": "Objective Verifiability",
             "CtxI": "Contextual Independence", "CM": "Capability Match"}
RATERS = ["Mehta", "Mukherjee"]          # Dash unavailable (workbook blank)

# ---------------- load ----------------
def load_ratings(who):
    from openpyxl import load_workbook
    ws = load_workbook(VAL / f"Ratings_{who}.xlsx", data_only=True)["Ratings"]
    rows = []
    for r in range(6, 106):
        rows.append({
            "task_id": int(ws.cell(row=r, column=2).value),
            f"{who}_IS": int(ws.cell(row=r, column=5).value),
            f"{who}_OV": int(ws.cell(row=r, column=6).value),
            f"{who}_CtxI": int(ws.cell(row=r, column=7).value),
            f"{who}_CM": int(ws.cell(row=r, column=8).value),
            f"{who}_flag": str(ws.cell(row=r, column=9).value or "").strip().upper(),
            f"{who}_note": str(ws.cell(row=r, column=10).value or "").strip(),
        })
    return pd.DataFrame(rows)

key = pd.read_csv(VAL / "validation_sample_100_KEY.csv")
key = key.rename(columns={
    "information_sufficiency_score": "LLM_IS", "objective_verifiability_score": "LLM_OV",
    "contextual_independence_score": "LLM_CtxI", "capability_match_score": "LLM_CM",
    "composite_score": "LLM_comp"})
d = key
for w in RATERS:
    d = d.merge(load_ratings(w), on="task_id")
assert len(d) == 100

flagged = d[(d["Mehta_flag"] == "Y") | (d["Mukherjee_flag"] == "Y")]["task_id"].tolist()
k = d[~d["task_id"].isin(flagged)].copy()     # kappa set, per protocol

# ---------------- stats ----------------
def kappa(a, b):
    n = len(a); po = sum(1 for x, y in zip(a, b) if x == y) / n
    ca, cb = Counter(a), Counter(b)
    pe = sum(ca[c] * cb.get(c, 0) for c in ca) / (n * n)
    return (po - pe) / (1 - pe) if pe < 1 else float("nan"), po, n

def kappa_linear(a, b):
    cats = sorted(set(a) | set(b)); n = len(a)
    O = defaultdict(float); ca, cb = Counter(a), Counter(b)
    for x, y in zip(a, b): O[(x, y)] += 1
    num = sum(abs(i - j) * O[(i, j)] for i in cats for j in cats)
    den = sum(abs(i - j) * ca[i] * cb[j] / n for i in cats for j in cats)
    return 1 - num / den if den > 0 else float("nan")

def kripp_alpha_ordinal(a, b):
    units = list(zip(a, b))
    o = defaultdict(float)
    for x, y in units:
        o[(x, y)] += 1; o[(y, x)] += 1
    n_c = defaultdict(float)
    for (c, _), v in o.items(): n_c[c] += v
    cats = sorted(n_c); n = sum(n_c.values())
    def d2(c, kk):
        if c == kk: return 0.0
        lo, hi = min(c, kk), max(c, kk)
        s = sum(n_c[g] for g in cats if lo <= g <= hi) - (n_c[lo] + n_c[hi]) / 2
        return s * s
    Do = sum(o[(c, kk)] * d2(c, kk) for c in cats for kk in cats) / n
    De = sum(n_c[c] * n_c[kk] * d2(c, kk) for c in cats for kk in cats) / (n * (n - 1))
    return 1 - Do / De if De > 0 else float("nan")

def design_y(is_, ov, ctxi, cm):
    if is_ == 0 or cm == 0: return 0.0
    return (is_ + ov + ctxi + cm) / 4 / 2

def confusion(a, b):
    cats = [0, 1, 2]
    m = pd.DataFrame(0, index=cats, columns=cats)
    for x, y in zip(a, b): m.loc[x, y] += 1
    return m

# ---------------- report ----------------
print("=" * 74)
print(f"TIER-2 HUMAN VALIDITY — {len(d)} tasks, raters: {', '.join(RATERS)} "
      f"(Dash workbook unfilled)")
print(f"Flagged-as-ambiguous (excluded from kappa): {len(flagged)} {flagged or ''}")
print("=" * 74)

print("\n[1] PER-CRITERION: LLM vs each rater  (kappa set n=%d)" % len(k))
print(f"    {'criterion':26s} {'rater':10s} {'%agree':>7s} {'kappa':>7s} {'lin-w kappa':>11s}")
pooled = {w: ([], []) for w in RATERS}
for c in CRIT:
    for w in RATERS:
        a = list(k[f"LLM_{c}"]); b = list(k[f"{w}_{c}"])
        kap, po, n = kappa(a, b); kw = kappa_linear(a, b)
        pooled[w][0].extend(a); pooled[w][1].extend(b)
        print(f"    {CRIT_FULL[c]:26s} {w:10s} {po*100:6.0f}% {kap:7.3f} {kw:11.3f}")

print("\n[2] POOLED ACROSS ALL FOUR CRITERIA (400 paired ratings per rater)")
for w in RATERS:
    kap, po, n = kappa(*pooled[w]); kw = kappa_linear(*pooled[w])
    print(f"    LLM vs {w:10s}: %agree={po*100:.0f}%  kappa={kap:.3f}  lin-w={kw:.3f}  (n={n})")

print("\n[3] CONSENSUS SUBSET (both raters agree) — cleanest human signal")
for c in CRIT:
    sub = k[k[f"{RATERS[0]}_{c}"] == k[f"{RATERS[1]}_{c}"]]
    a = list(sub[f"LLM_{c}"]); b = list(sub[f"{RATERS[0]}_{c}"])
    if len(sub) >= 10:
        kap, po, n = kappa(a, b)
        print(f"    {CRIT_FULL[c]:26s}: n={n:3d}  %agree={po*100:.0f}%  kappa={kap:.3f}")

print("\n[4] INTER-RATER (human vs human)")
ir_pool = ([], [])
for c in CRIT:
    a = list(k[f"{RATERS[0]}_{c}"]); b = list(k[f"{RATERS[1]}_{c}"])
    ir_pool[0].extend(a); ir_pool[1].extend(b)
    kap, po, n = kappa(a, b); al = kripp_alpha_ordinal(a, b)
    print(f"    {CRIT_FULL[c]:26s}: %agree={po*100:.0f}%  kappa={kap:.3f}  Kripp-alpha(ord)={al:.3f}")
kap, po, n = kappa(*ir_pool)
print(f"    {'POOLED':26s}: %agree={po*100:.0f}%  kappa={kap:.3f}  "
      f"Kripp-alpha(ord)={kripp_alpha_ordinal(*ir_pool):.3f}")

print("\n[5] COMPOSITE (Design Y from each rater's scores) vs LLM composite")
for w in RATERS:
    comp = [design_y(r[f"{w}_IS"], r[f"{w}_OV"], r[f"{w}_CtxI"], r[f"{w}_CM"])
            for _, r in k.iterrows()]
    llm = list(k["LLM_comp"].astype(float))
    gate_h = [1 if c == 0 else 0 for c in comp]; gate_l = [1 if c == 0 else 0 for c in llm]
    gk, gpo, _ = kappa(gate_h, gate_l)
    rho = pd.Series(comp).corr(pd.Series(llm), method="spearman")
    mae = sum(abs(a - b) for a, b in zip(comp, llm)) / len(comp)
    print(f"    {w:10s}: gate agreement={gpo*100:.0f}% (kappa={gk:.3f})  "
          f"Spearman rho={rho:.3f}  MAE={mae:.3f}")

print("\n[6] CONFUSION MATRICES (rows=LLM, cols=rater) — per criterion, per rater")
for c in CRIT:
    for w in RATERS:
        m = confusion(list(k[f"LLM_{c}"]), list(k[f"{w}_{c}"]))
        flat = "  ".join(f"{m.loc[i, j]:3d}" for i in [0, 1, 2] for j in [0, 1, 2])
        print(f"    {c:5s} vs {w:10s}: [r0: {flat[:13]}] [r1: {flat[14:27]}] [r2: {flat[28:]}]")

print("\n[7] LARGEST DISAGREEMENTS (|LLM - rater-mean| >= 1.5 on any criterion)")
big = []
for _, r in k.iterrows():
    for c in CRIT:
        hm = (r[f"{RATERS[0]}_{c}"] + r[f"{RATERS[1]}_{c}"]) / 2
        gap = r[f"LLM_{c}"] - hm
        if abs(gap) >= 1.5:
            big.append((r["task_id"], c, r[f"LLM_{c}"], hm, gap, r["occupation_title"], r["task_text"]))
print(f"    {len(big)} criterion-level gaps >= 1.5")
for tid, c, l, h, g, occ, txt in big:
    print(f"    [{c}] LLM={l} humans={h:.1f}  {occ[:30]}: {str(txt)[:60]}")

print("\n[8] DIRECTION OF DISAGREEMENT (LLM minus rater, mean signed gap)")
for c in CRIT:
    for w in RATERS:
        gaps = (k[f"LLM_{c}"] - k[f"{w}_{c}"])
        print(f"    {c:5s} vs {w:10s}: mean={gaps.mean():+.3f}   "
              f"LLM higher on {int((gaps>0).sum())}, lower on {int((gaps<0).sum())}, equal {int((gaps==0).sum())}")

print("\n[9] SAMPLE SCORE DISTRIBUTION (prevalence context for kappa)")
for c in CRIT:
    print(f"    {c:5s}: LLM {dict(Counter(k['LLM_'+c]))}   "
          + "   ".join(f"{w} {dict(Counter(k[w+'_'+c]))}" for w in RATERS))
print(f"    LLM gate rate in sample: {(k['LLM_comp']==0).mean()*100:.0f}%")
