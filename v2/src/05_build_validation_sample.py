"""
05_build_validation_sample.py
=============================
Draw the Tier-2 human-validity sample and build the blind rating workbooks.

- Simple random sample of N=100 tasks (fixed seed) from the scored pre-run pool
  (scored_master_tasks_shuffled.csv). Simple random, not stratified, per
  Decision History 5.27.
- Writes a researcher-only ANSWER KEY (task_id + the LLM's per-criterion scores
  and composite) for computing κ later — NOT given to raters.
- Writes one BLIND workbook per rater: task_id, occupation, task text, and four
  empty 0/1/2 score columns (dropdown-validated), plus Flag/Notes. No LLM scores.
  Each workbook also carries a "Read me first" sheet whose criterion definitions
  and anchors are pulled VERBATIM from rubric_prompt.py (the same words the model
  saw), so raters and the model judge against identical wording.

Reproducible: fixed seed. Author: Aaryan Mehta. Created 06 June 2026.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import re
import importlib.util
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
SCORED = ROOT / "data" / "processed" / "scored_master_tasks_shuffled.csv"
OUTDIR = ROOT / "data" / "processed" / "validation"
OUTDIR.mkdir(parents=True, exist_ok=True)

SAMPLE_SEED = 20260606
N = 100
RATERS = ["Mehta", "Dash", "Mukherjee"]
CRIT = ["Information Sufficiency", "Objective Verifiability",
        "Contextual Independence", "Capability Match"]

# --- verbatim criterion text from the locked prompt -----------------------
spec = importlib.util.spec_from_file_location("rp", ROOT / "src" / "rubric_prompt.py")
rp = importlib.util.module_from_spec(spec); spec.loader.exec_module(rp)
SP = rp.SYSTEM_PROMPT
CONSTRUCT = re.search(r"(You are evaluating whether.+?proficiency\.)", SP, re.S).group(1)
DISAMBIG = re.search(r"(When a task could plausibly.+?the other three criteria\.)", SP, re.S).group(1)

# Parse each criterion block into (name, question, guard, anchors), verbatim.
# The guard is the text between the question and the first anchor, with the
# model-only output mechanics (the two reasoning-field instructions) filtered out.
CRIT_BLOCKS = []
for blk in re.split(r"## Criterion \d — ", SP)[1:]:
    blk = blk.split("# Output")[0]
    head = blk.partition("\n- 0:")[0]
    lines = [l.strip() for l in head.split("\n") if l.strip()]
    name = lines[0]
    body = [l for l in lines[1:]
            if not (l.startswith("For this criterion") or "capability_match_what" in l
                    or l.startswith("Both fields are required"))]
    question = body[0]
    guard = " ".join(body[1:])
    anchors = re.findall(r"- ([012]: .+)", blk)
    CRIT_BLOCKS.append((name, question, guard, anchors))

# --- draw the sample ------------------------------------------------------
d = pd.read_csv(SCORED)
d = d[d["scoring_status"] != "TERMINAL_FAILURE"].copy()
sample = d.sample(n=N, random_state=SAMPLE_SEED).sort_values("task_id").reset_index(drop=True)

# --- researcher-only answer key ------------------------------------------
key_cols = ["task_id", "occupation_title", "task_text",
            "information_sufficiency_score", "objective_verifiability_score",
            "contextual_independence_score", "capability_match_score",
            "composite_score", "scoring_status"]
key_path = OUTDIR / "validation_sample_100_KEY.csv"
sample[key_cols].to_csv(key_path, index=False, encoding="utf-8")

# --- styling helpers ------------------------------------------------------
NAVY = "1E3A5F"; HFILL = PatternFill("solid", fgColor=NAVY)
SOFT = PatternFill("solid", fgColor="EEF3F8")
SCORE_FILL = PatternFill("solid", fgColor="FFF7E6")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def build_blind(rater):
    wb = Workbook()

    # ---- Sheet 1: Read me first (verbatim criteria) ----
    ins = wb.active; ins.title = "Read me first"
    ins.column_dimensions["A"].width = 110
    rows = [
        ("Tier-2 Human Validity Ratings — GenAI Exposure Index", 16, True, NAVY),
        ("Score each of the 100 tasks (on the 'Ratings' tab) on the four criteria below, each 0 / 1 / 2. "
         "Work BLIND — do not look at anyone else's scores or any model output. If a task is genuinely too "
         "ambiguous to score, put Y in the Flag column and a note. The full guide with worked examples is "
         "Rater_Guide.docx; the definitions below are reproduced verbatim from the instructions the model was given.", 11, False, "000000"),
        ("", 8, False, "000000"),
        ("What you are evaluating", 12, True, NAVY),
        (CONSTRUCT, 11, False, "000000"),
        ("", 8, False, "000000"),
        ("Disambiguation rule", 12, True, NAVY),
        (DISAMBIG, 11, False, "000000"),
        ("", 8, False, "000000"),
    ]
    for i, (name, question, guard, anchors) in enumerate(CRIT_BLOCKS):
        rows.append((f"Criterion {i+1} — {name}", 12, True, NAVY))
        rows.append((question, 11, False, "000000"))
        if guard:
            rows.append((guard, 10, False, "595959"))
        for a in anchors:
            rows.append(("    " + a, 11, False, "000000"))
        rows.append(("", 8, False, "000000"))
    r = 1
    for text, size, bold, color in rows:
        c = ins.cell(row=r, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = WRAP
        r += 1

    # ---- Sheet 2: Ratings (blind) ----
    ws = wb.create_sheet("Ratings")
    ws["A1"] = "Tier-2 Human Validity Ratings — GenAI Exposure Index"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = f"Rater: {rater}"; ws["A2"].font = BOLD
    ws["C2"] = "Date: ____________"
    ws["A3"] = ("Score each criterion 0 / 1 / 2 (dropdowns). Blind to others and to any model output. "
                "Flag = Y if too ambiguous. Do not edit task_id / Occupation / Task.")
    ws["A3"].font = Font(italic=True, color="595959")

    headers = ["No.", "task_id", "Occupation", "Task"] + CRIT + ["Flag (Y/N)", "Notes"]
    HROW = 5
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=HROW, column=j, value=h)
        c.font = WHITE_BOLD; c.fill = HFILL; c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = BORDER

    for i, row in sample.iterrows():
        rr = HROW + 1 + i
        vals = [i + 1, int(row["task_id"]), row["occupation_title"], row["task_text"], "", "", "", "", "", ""]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=rr, column=j, value=v)
            c.border = BORDER
            c.alignment = WRAP if j in (3, 4, 10) else CENTER
            if j == 1 or j == 2:
                c.alignment = CENTER
            if 5 <= j <= 8:
                c.fill = SCORE_FILL
            elif rr % 2 == 0:
                c.fill = SOFT

    # widths
    widths = {"A": 5, "B": 9, "C": 30, "D": 78, "E": 15, "F": 15, "G": 15, "H": 15, "I": 11, "J": 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[HROW].height = 42

    # data validation: scores 0/1/2 ; flag Y/N
    last = HROW + N
    dv_score = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True,
                              showErrorMessage=True, errorTitle="Invalid", error="Enter 0, 1, or 2.")
    ws.add_data_validation(dv_score); dv_score.add(f"E{HROW+1}:H{last}")
    dv_flag = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_flag); dv_flag.add(f"I{HROW+1}:I{last}")

    ws.freeze_panes = "E6"   # keep No./task_id/Occupation/Task + header visible

    out = OUTDIR / f"Ratings_{rater}.xlsx"
    wb.save(out)
    return out


paths = [build_blind(r) for r in RATERS]
print(f"Sample: {N} tasks (simple random, seed={SAMPLE_SEED}) from {len(d)} scored tasks.")
print(f"  sample gate rate: {(sample['scoring_status']=='GATED').mean()*100:.0f}%  "
      f"(pool {(d['scoring_status']=='GATED').mean()*100:.0f}%)")
print(f"Answer key (researcher-only): {key_path}")
for p in paths:
    print(f"Blind workbook: {p}")
